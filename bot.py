import os
import re
import asyncio
import tempfile
from openpyxl import load_workbook
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters,
)

BOT_TOKEN = os.getenv("BOT_TOKEN")

keyboard = ReplyKeyboardMarkup(
    [["Zvonok", "Gudman"]],
    resize_keyboard=True
)

user_queues = {}
user_tasks = {}

# =========================
# PHONE CLEANER
# =========================
def clean_phone(value):
    if value is None:
        return None

    digits = re.sub(r"\D", "", str(value))

    if len(digits) == 11 and digits.startswith("8"):
        digits = "7" + digits[1:]

    if len(digits) == 10:
        digits = "7" + digits

    return digits if len(digits) >= 10 else None


# =========================
# GUDMAN (FAST)
# =========================
def process_gudman(file_path):
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    return [
        phone
        for row in ws.iter_rows(values_only=True)
        if (phone := clean_phone(row[0] if row else None))
    ]


# =========================
# ZVONOK (OPTIMIZED)
# =========================
def process_zvonok(file_path):
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    headers = next(rows)

    headers_lower = [(h or "").strip().lower() for h in headers]

    phone_col = status_col = client_answer_col = None

    for i, h in enumerate(headers_lower):
        if "номер" in h and "телефон" in h:
            phone_col = i
        elif "статус звонка" in h:
            status_col = i
        elif (
            "транскрибация клиента" in h
            or "транскрипция клиента" in h
            or "ответ клиента" in h
        ):
            client_answer_col = i

    if phone_col is None:
        raise Exception("Не нашёл столбец 'Номер телефона'.")
    if status_col is None:
        raise Exception("Не нашёл столбец 'Статус звонка'.")
    if client_answer_col is None:
        raise Exception("Не нашёл столбец 'Транскрибация клиента' или 'Ответ клиента'.")

    bad_answers = {
        "номер набран неправильно",
        "проверьте корректность наборам",
        "проверьте корректность набора",
        "номер не используется",
    }

    blocked_m_text = "набранный вами номер не"

    numbers = []

    for row in rows:
        if not row:
            continue

        status = (row[status_col] or "").strip().lower()
        if status != "закончен удачно":
            continue

        client_answer = (row[client_answer_col] or "").strip().lower()

        # колонка M (индекс 12)
        col_m = (row[12] or "").strip().lower() if len(row) > 12 else ""

        if blocked_m_text in col_m:
            continue

        if any(bad in client_answer for bad in bad_answers):
            continue

        phone = clean_phone(row[phone_col])
        if phone:
            numbers.append(phone)

    return numbers


# =========================
# START
# =========================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["mode"] = None

    await update.message.reply_text(
        "Выбери режим обработки:",
        reply_markup=keyboard
    )


# =========================
# MODE SELECT
# =========================
async def choose_mode(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip().lower()

    if text == "zvonok":
        context.user_data["mode"] = "Zvonok"
    elif text == "gudman":
        context.user_data["mode"] = "Gudman"
    else:
        await update.message.reply_text(
            "Нажми кнопку Zvonok или Gudman.",
            reply_markup=keyboard
        )
        return

    await update.message.reply_text(
        f"Режим выбран: {context.user_data['mode']} ✅\n"
        f"Кидай Excel-файлы .xlsx"
    )


# =========================
# FILE HANDLER
# =========================
async def handle_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    mode = context.user_data.get("mode")

    if not mode:
        await update.message.reply_text(
            "Сначала выбери режим: Zvonok или Gudman.",
            reply_markup=keyboard
        )
        return

    document = update.message.document

    if not document.file_name.lower().endswith(".xlsx"):
        await update.message.reply_text("Отправь Excel-файл в формате .xlsx")
        return

    user_id = update.effective_user.id

    user_queues.setdefault(user_id, []).append({
        "document": document,
        "mode": mode,
        "chat_id": update.effective_chat.id
    })

    await update.message.reply_text(f"Файл принят ✅\n{document.file_name}")

    if user_id not in user_tasks or user_tasks[user_id].done():
        user_tasks[user_id] = asyncio.create_task(process_queue(context, user_id))


# =========================
# QUEUE PROCESSOR (FAST)
# =========================
async def process_queue(context: ContextTypes.DEFAULT_TYPE, user_id: int):
    await asyncio.sleep(1)  # было 3 → ускорили

    queue = user_queues.get(user_id, [])
    if not queue:
        return

    total = len(queue)
    chat_id = queue[0]["chat_id"]

    progress_message = await context.bot.send_message(
        chat_id=chat_id,
        text=f"🔴 0% | Начинаю обработку файлов: {total}"
    )

    processed = 0

    while user_queues.get(user_id):
        item = user_queues[user_id].pop(0)

        document = item["document"]
        mode = item["mode"]

        input_path = None
        output_path = None

        try:
            percent = int((processed / total) * 100)
            emoji = get_progress_emoji(percent)

            await progress_message.edit_text(
                f"{emoji} {percent}% | Файл {processed + 1} из {total}\n"
                f"{document.file_name}"
            )

            tg_file = await document.get_file()

            safe_name = document.file_name.replace("/", "_").replace("\\", "_")
            input_path = os.path.join(tempfile.gettempdir(), safe_name)

            await tg_file.download_to_drive(input_path)

            # 🔥 ВАЖНО: CPU-heavy в отдельный поток
            if mode == "Zvonok":
                numbers = await asyncio.to_thread(process_zvonok, input_path)
                output_name = safe_name.replace(".xlsx", "_zvonok.txt")
            else:
                numbers = await asyncio.to_thread(process_gudman, input_path)
                output_name = safe_name.replace(".xlsx", "_gudman.txt")

            output_path = os.path.join(tempfile.gettempdir(), output_name)

            with open(output_path, "w", encoding="utf-8") as f:
                f.write("\n".join(numbers))

            with open(output_path, "rb") as file:
                await context.bot.send_document(
                    chat_id=chat_id,
                    document=file,
                    filename=output_name,
                    caption=(
                        f"Готово ✅\n"
                        f"Файл: {document.file_name}\n"
                        f"Режим: {mode}\n"
                        f"Найдено номеров: {len(numbers)}"
                    )
                )

        except Exception as e:
            await context.bot.send_message(
                chat_id=chat_id,
                text=f"Ошибка при обработке файла {document.file_name}:\n{e}"
            )

        finally:
            if input_path and os.path.exists(input_path):
                os.remove(input_path)
            if output_path and os.path.exists(output_path):
                os.remove(output_path)

        processed += 1

    await progress_message.edit_text("✅ 100% | Все файлы обработаны")


# =========================
# PROGRESS EMOJI
# =========================
def get_progress_emoji(percent):
    if percent < 25:
        return "🔴"
    elif percent < 50:
        return "🟠"
    elif percent < 75:
        return "🟡"
    elif percent < 100:
        return "🟢"
    return "✅"


# =========================
# MAIN
# =========================
def main():
    if not BOT_TOKEN:
        raise ValueError("BOT_TOKEN не найден")

    app = ApplicationBuilder().token(BOT_TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, choose_mode))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_file))

    print("🚀 Бот запущен")
    app.run_polling()


if __name__ == "__main__":
    main()
